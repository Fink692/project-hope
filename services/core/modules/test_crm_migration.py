import csv
import io
import json
import zipfile
from datetime import timedelta

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient

from audit.models import AuditEvent
from identity.models import Membership, Organization, User

from .models import (
    ConsentRecord,
    Contact,
    ContactRelationship,
    DonorSnapshot,
    EmailMessage,
    Household,
    Interaction,
    Mailbox,
    Program,
    RetentionPolicy,
    ScheduleEvent,
    VolunteerProfile,
    WaitlistEntry,
)


class CRMMigrationApiTests(TestCase):
    password = "Migration-password-123"

    def setUp(self):
        self.client = APIClient()
        self.owner = User.objects.create_user("owner@example.org", self.password)
        self.second_admin = User.objects.create_user("admin@example.org", self.password)
        self.viewer = User.objects.create_user("viewer@example.org", self.password)
        self.foreign_owner = User.objects.create_user(
            "foreign@example.org", self.password
        )
        self.organization = Organization.objects.create(
            name="Migration Charity", slug="migration-charity"
        )
        self.foreign_organization = Organization.objects.create(
            name="Foreign Charity", slug="foreign-charity"
        )
        Membership.objects.create(
            organization=self.organization,
            user=self.owner,
            role=Membership.Role.OWNER,
        )
        Membership.objects.create(
            organization=self.organization,
            user=self.second_admin,
            role=Membership.Role.ADMIN,
        )
        Membership.objects.create(
            organization=self.organization,
            user=self.viewer,
            role=Membership.Role.VIEWER,
        )
        Membership.objects.create(
            organization=self.foreign_organization,
            user=self.foreign_owner,
            role=Membership.Role.OWNER,
        )
        self.assertTrue(
            self.client.login(email=self.owner.email, password=self.password)
        )

    def csv_upload(self, rows, name="contacts.csv"):
        output = io.StringIO(newline="")
        writer = csv.writer(output, lineterminator="\r\n")
        writer.writerows(rows)
        return SimpleUploadedFile(
            name, output.getvalue().encode("utf-8"), content_type="text/csv"
        )

    def xlsx_upload(self, rows, name="contacts.xlsx"):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Contacts"
        for row in rows:
            worksheet.append(row)
        output = io.BytesIO()
        workbook.save(output)
        workbook.close()
        return SimpleUploadedFile(
            name,
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def preview(self, upload):
        return self.client.post(
            "/api/v1/organizations/migration-charity/crm/imports/preview/",
            {"file": upload},
            format="multipart",
        )

    def commit(self, upload, token, actions):
        return self.client.post(
            "/api/v1/organizations/migration-charity/crm/imports/commit/",
            {
                "file": upload,
                "previewToken": token,
                "actions": json.dumps(actions),
            },
            format="multipart",
        )

    def test_csv_preview_and_reviewed_import_fill_blanks_without_overwrite(self):
        existing = Contact.objects.create(
            organization=self.organization,
            first_name="Amina",
            last_name="Hope",
            email="amina@example.org",
            notes="Reviewed existing note",
        )
        rows = [
            ["First Name", "Last Name", "Email Address", "Phone Number", "Notes"],
            [
                "Different",
                "Name",
                "AMINA@example.org",
                "+1 204 555 0100",
                "Imported context",
            ],
            ["Sam", "Volunteer", "sam@example.org", "", "New contact"],
        ]
        preview = self.preview(self.csv_upload(rows))
        self.assertEqual(preview.status_code, 200, preview.content)
        body = preview.json()
        self.assertEqual(body["summary"]["exactMatches"], 1)
        self.assertEqual(body["summary"]["newRecords"], 1)
        exact = body["rows"][0]
        self.assertEqual(exact["candidates"][0]["id"], str(existing.id))

        committed = self.commit(
            self.csv_upload(rows),
            body["previewToken"],
            [
                {
                    "rowNumber": 2,
                    "action": "update",
                    "targetContactId": str(existing.id),
                },
                {"rowNumber": 3, "action": "create"},
            ],
        )
        self.assertEqual(committed.status_code, 201, committed.content)
        self.assertEqual(committed.json()["created"], 1)
        self.assertEqual(committed.json()["updated"], 1)
        existing.refresh_from_db()
        self.assertEqual(existing.first_name, "Amina")
        self.assertEqual(existing.last_name, "Hope")
        self.assertEqual(existing.phone, "+1 204 555 0100")
        self.assertIn("Reviewed existing note", existing.notes)
        self.assertIn("Imported context", existing.notes)
        self.assertTrue(
            Contact.objects.filter(
                organization=self.organization, email="sam@example.org"
            ).exists()
        )
        event = AuditEvent.objects.filter(action="contact.import_committed").latest(
            "created_at"
        )
        self.assertNotIn("amina@example.org", json.dumps(event.metadata))

    def test_preview_is_short_lived_user_tenant_and_file_bound(self):
        rows = [["First Name", "Email"], ["Amina", "amina@example.org"]]
        preview = self.preview(self.csv_upload(rows)).json()

        changed = [["First Name", "Email"], ["Changed", "changed@example.org"]]
        wrong_file = self.commit(
            self.csv_upload(changed),
            preview["previewToken"],
            [{"rowNumber": 2, "action": "create"}],
        )
        self.assertEqual(wrong_file.status_code, 409)
        self.assertEqual(Contact.objects.count(), 0)

        self.client.logout()
        self.assertTrue(
            self.client.login(email=self.second_admin.email, password=self.password)
        )
        wrong_user = self.commit(
            self.csv_upload(rows),
            preview["previewToken"],
            [{"rowNumber": 2, "action": "create"}],
        )
        self.assertEqual(wrong_user.status_code, 409)
        self.assertEqual(Contact.objects.count(), 0)

        Membership.objects.create(
            organization=self.foreign_organization,
            user=self.second_admin,
            role=Membership.Role.ADMIN,
        )
        cross_tenant = self.client.post(
            "/api/v1/organizations/foreign-charity/crm/imports/commit/",
            {
                "file": self.csv_upload(rows),
                "previewToken": preview["previewToken"],
                "actions": json.dumps([{"rowNumber": 2, "action": "create"}]),
            },
            format="multipart",
        )
        self.assertEqual(cross_tenant.status_code, 409)
        self.assertEqual(Contact.objects.count(), 0)

    @override_settings(PROJECT_HOPE_CRM_IMPORT_PREVIEW_MAX_AGE_SECONDS=-1)
    def test_expired_preview_cannot_commit(self):
        rows = [["First Name", "Email"], ["Amina", "amina@example.org"]]
        preview = self.preview(self.csv_upload(rows)).json()
        response = self.commit(
            self.csv_upload(rows),
            preview["previewToken"],
            [{"rowNumber": 2, "action": "create"}],
        )
        self.assertEqual(response.status_code, 409)
        self.assertIn("expired", response.json()["detail"].lower())

    def test_xlsx_formula_is_rejected_at_row_level(self):
        rows = [
            ["First Name", "Email", "Notes"],
            ["Amina", "amina@example.org", '=HYPERLINK("https://invalid")'],
        ]
        preview = self.preview(self.xlsx_upload(rows))
        self.assertEqual(preview.status_code, 200, preview.content)
        row = preview.json()["rows"][0]
        self.assertEqual(row["status"], "invalid")
        self.assertIn("formulas", row["errors"]["notes"][0].lower())
        committed = self.commit(
            self.xlsx_upload(rows),
            preview.json()["previewToken"],
            [{"rowNumber": 2, "action": "skip"}],
        )
        self.assertEqual(committed.status_code, 201, committed.content)
        self.assertEqual(committed.json()["invalid"], 1)
        self.assertEqual(committed.json()["skipped"], 0)

    def test_malformed_legacy_and_macro_workbooks_fail_closed(self):
        malformed = self.preview(
            SimpleUploadedFile(
                "broken.xlsx",
                b"not an office archive",
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        )
        self.assertEqual(malformed.status_code, 400)

        legacy = self.preview(
            SimpleUploadedFile(
                "legacy.xls",
                b"legacy workbook",
                content_type="application/vnd.ms-excel",
            )
        )
        self.assertEqual(legacy.status_code, 415)
        self.assertIn("saved as xlsx", legacy.json()["detail"].lower())

        valid_workbook = self.xlsx_upload(
            [["First Name", "Email"], ["Amina", "amina@example.org"]]
        ).read()
        macro_workbook = io.BytesIO(valid_workbook)
        with zipfile.ZipFile(macro_workbook, "a") as archive:
            archive.writestr("xl/vbaProject.bin", b"not a real macro")
        macro = self.preview(
            SimpleUploadedFile(
                "macro.xlsx",
                macro_workbook.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        )
        self.assertEqual(macro.status_code, 400)
        self.assertIn("macro", macro.json()["detail"].lower())
        self.assertEqual(Contact.objects.count(), 0)

    @override_settings(PROJECT_HOPE_MAX_CRM_IMPORT_ROWS=1)
    def test_import_row_limit_fails_closed(self):
        response = self.preview(
            self.csv_upload(
                [
                    ["First Name", "Email"],
                    ["One", "one@example.org"],
                    ["Two", "two@example.org"],
                ]
            )
        )
        self.assertEqual(response.status_code, 413)
        self.assertEqual(Contact.objects.count(), 0)

    def test_duplicate_rows_in_same_file_are_invalid(self):
        preview = self.preview(
            self.csv_upload(
                [
                    ["First Name", "Email", "External Reference"],
                    ["One", "same@example.org", "legacy-1"],
                    ["Two", "SAME@example.org", "legacy-2"],
                ]
            )
        ).json()
        self.assertEqual(preview["rows"][0]["status"], "new")
        self.assertEqual(preview["rows"][1]["status"], "invalid")
        self.assertIn("row 2", preview["rows"][1]["errors"]["email"][0].lower())

    def test_template_and_exports_are_safe_and_round_trip(self):
        Contact.objects.create(
            organization=self.organization,
            first_name="Formula",
            email="formula@example.org",
            notes='=HYPERLINK("https://malicious.invalid","open")',
        )
        template = self.client.get(
            "/api/v1/organizations/migration-charity/crm/template/?fileFormat=xlsx"
        )
        self.assertEqual(template.status_code, 200)
        template_book = load_workbook(io.BytesIO(template.content), data_only=False)
        self.assertEqual(template_book.sheetnames, ["Contacts", "Read me"])
        self.assertEqual(template_book["Contacts"]["A1"].value, "contact_type")
        template_book.close()

        xlsx = self.client.get(
            "/api/v1/organizations/migration-charity/crm/export/?fileFormat=xlsx"
        )
        self.assertEqual(xlsx.status_code, 200)
        exported_book = load_workbook(io.BytesIO(xlsx.content), data_only=False)
        contact_sheet = exported_book["Contacts"]
        notes_column = [cell.value for cell in contact_sheet[1]].index("notes") + 1
        exported_note = contact_sheet.cell(2, notes_column)
        self.assertEqual(exported_note.data_type, "s")
        self.assertTrue(exported_note.value.startswith("="))
        exported_book.close()

        csv_response = self.client.get(
            "/api/v1/organizations/migration-charity/crm/export/?fileFormat=csv"
        )
        self.assertEqual(csv_response.status_code, 200)
        self.assertTrue(csv_response.content.startswith(b"\xef\xbb\xbf"))
        self.assertIn(b"'=HYPERLINK", csv_response.content)

        round_trip = self.preview(
            SimpleUploadedFile(
                "round-trip.csv", csv_response.content, content_type="text/csv"
            )
        )
        self.assertEqual(round_trip.status_code, 200, round_trip.content)
        self.assertEqual(round_trip.json()["summary"]["exactMatches"], 1)
        self.assertIn("export-only", " ".join(round_trip.json()["warnings"]))
        self.assertEqual(round_trip["Cache-Control"], "private, no-store")
        self.assertEqual(xlsx["Cache-Control"], "private, no-store")

    def test_viewer_can_read_but_cannot_write_or_bulk_export(self):
        contact = Contact.objects.create(
            organization=self.organization, first_name="Existing"
        )
        self.client.logout()
        self.assertTrue(
            self.client.login(email=self.viewer.email, password=self.password)
        )
        listing = self.client.get("/api/v1/organizations/migration-charity/contacts/")
        self.assertEqual(listing.status_code, 200)
        create = self.client.post(
            "/api/v1/organizations/migration-charity/contacts/",
            {"first_name": "Not allowed"},
            format="json",
        )
        self.assertEqual(create.status_code, 403)
        update = self.client.patch(
            f"/api/v1/organizations/migration-charity/contacts/{contact.id}/",
            {"first_name": "Not allowed"},
            format="json",
        )
        self.assertEqual(update.status_code, 403)
        contact.refresh_from_db()
        self.assertEqual(contact.first_name, "Existing")
        preview = self.preview(self.csv_upload([["First Name"], ["Not allowed"]]))
        self.assertEqual(preview.status_code, 403)
        export = self.client.get("/api/v1/organizations/migration-charity/crm/export/")
        self.assertEqual(export.status_code, 403)

    def test_contact_editor_requires_a_usable_identity(self):
        endpoint = "/api/v1/organizations/migration-charity/contacts/"
        empty = self.client.post(endpoint, {}, format="json")
        self.assertEqual(empty.status_code, 400)
        self.assertIn("Provide a name", str(empty.json()))

        unnamed_organization = self.client.post(
            endpoint,
            {"contact_type": "organization", "email": "office@example.org"},
            format="json",
        )
        self.assertEqual(unnamed_organization.status_code, 400)
        self.assertIn("organization_name", unnamed_organization.json())

        preferred_only = self.client.post(
            endpoint,
            {"preferred_name": "River"},
            format="json",
        )
        self.assertEqual(preferred_only.status_code, 201, preferred_only.content)
        cleared = self.client.patch(
            f"{endpoint}{preferred_only.json()['id']}/",
            {"preferred_name": ""},
            format="json",
        )
        self.assertEqual(cleared.status_code, 400)

        preview = self.preview(
            self.csv_upload([["Preferred name"], ["Sky"]], "preferred-name.csv")
        )
        self.assertEqual(preview.status_code, 200, preview.content)
        self.assertEqual(preview.json()["summary"]["newRecords"], 1)

    def test_duplicate_review_and_merge_preserve_source_and_related_records(self):
        primary = Contact.objects.create(
            organization=self.organization,
            first_name="Amina",
            email="same@example.org",
            sensitivity="internal",
            consent_status="unknown",
            notes="Primary note",
        )
        duplicate = Contact.objects.create(
            organization=self.organization,
            last_name="Hope",
            email="same@example.org",
            phone="204-555-0100",
            sensitivity="restricted",
            consent_status="withdrawn",
            notes="Duplicate note",
        )
        other = Contact.objects.create(
            organization=self.organization, first_name="Related"
        )
        cross_type = Contact.objects.create(
            organization=self.organization,
            contact_type="organization",
            organization_name="Shared Inbox Organization",
            email="same@example.org",
        )
        household = Household.objects.create(
            organization=self.organization,
            name="Hope household",
            primary_contact=duplicate,
        )
        interaction = Interaction.objects.create(
            organization=self.organization,
            contact=duplicate,
            body="Called",
            occurred_at=timezone.now(),
        )
        inconsistent_foreign_interaction = Interaction.objects.create(
            organization=self.foreign_organization,
            contact=duplicate,
            body="Foreign tenant record",
            occurred_at=timezone.now(),
        )
        consent = ConsentRecord.objects.create(
            organization=self.organization,
            contact=duplicate,
            purpose="Updates",
            status="withdrawn",
        )
        program = Program.objects.create(
            organization=self.organization, name="Community programme"
        )
        schedule = ScheduleEvent.objects.create(
            organization=self.organization,
            title="Appointment",
            starts_at=timezone.now(),
            ends_at=timezone.now() + timedelta(hours=1),
            contact=duplicate,
        )
        waitlist = WaitlistEntry.objects.create(
            organization=self.organization,
            contact=duplicate,
            program=program,
            requested_at=timezone.now(),
        )
        mailbox = Mailbox.objects.create(
            organization=self.organization,
            name="Inbox",
            address="inbox@example.org",
        )
        message = EmailMessage.objects.create(
            organization=self.organization,
            mailbox=mailbox,
            external_id="message-1",
            sender="sender@example.org",
            subject="Hello",
            body_excerpt="Hello",
            received_at=timezone.now(),
            crm_contact=duplicate,
        )
        donor = DonorSnapshot.objects.create(
            organization=self.organization,
            contact=duplicate,
            period_start="2026-01-01",
            period_end="2026-06-30",
        )
        volunteer = VolunteerProfile.objects.create(
            organization=self.organization, contact=duplicate
        )
        relationship = ContactRelationship.objects.create(
            organization=self.organization,
            from_contact=duplicate,
            to_contact=other,
            relation_type="family",
            notes="Sibling",
        )

        candidates = self.client.get(
            "/api/v1/organizations/migration-charity/crm/duplicates/"
        )
        self.assertEqual(candidates.status_code, 200)
        self.assertEqual(candidates.json()["totalCandidates"], 1)
        self.assertEqual(candidates.json()["results"][0]["confidence"], "exact")
        candidate_ids = {
            candidates.json()["results"][0]["first"]["id"],
            candidates.json()["results"][0]["second"]["id"],
        }
        self.assertNotIn(str(cross_type.id), candidate_ids)

        unrelated = self.client.post(
            "/api/v1/organizations/migration-charity/crm/duplicates/merge/",
            {
                "primaryContactId": str(primary.id),
                "duplicateContactId": str(other.id),
                "confirm": True,
            },
            format="json",
        )
        self.assertEqual(unrelated.status_code, 409)

        merged = self.client.post(
            "/api/v1/organizations/migration-charity/crm/duplicates/merge/",
            {
                "primaryContactId": str(primary.id),
                "duplicateContactId": str(duplicate.id),
                "confirm": True,
            },
            format="json",
        )
        self.assertEqual(merged.status_code, 200, merged.content)
        primary.refresh_from_db()
        duplicate.refresh_from_db()
        self.assertEqual(primary.last_name, "Hope")
        self.assertEqual(primary.phone, "204-555-0100")
        self.assertEqual(primary.sensitivity, "restricted")
        self.assertEqual(primary.consent_status, "withdrawn")
        self.assertIn("Primary note", primary.notes)
        self.assertIn("Duplicate note", primary.notes)
        self.assertEqual(duplicate.record_status, Contact.RecordStatus.MERGED)
        self.assertEqual(duplicate.merged_into, primary)
        self.assertTrue(Contact.objects.filter(pk=duplicate.id).exists())
        for related in (
            household,
            interaction,
            consent,
            schedule,
            waitlist,
            message,
            donor,
            volunteer,
            relationship,
        ):
            related.refresh_from_db()
        self.assertEqual(household.primary_contact, primary)
        self.assertEqual(interaction.contact, primary)
        inconsistent_foreign_interaction.refresh_from_db()
        self.assertEqual(inconsistent_foreign_interaction.contact, duplicate)
        self.assertEqual(consent.contact, primary)
        self.assertEqual(schedule.contact, primary)
        self.assertEqual(waitlist.contact, primary)
        self.assertEqual(message.crm_contact, primary)
        self.assertEqual(donor.contact, primary)
        self.assertEqual(volunteer.contact, primary)
        self.assertEqual(relationship.from_contact, primary)
        listing = self.client.get(
            "/api/v1/organizations/migration-charity/contacts/"
        ).json()
        self.assertNotIn(str(duplicate.id), {item["id"] for item in listing})
        event = AuditEvent.objects.filter(action="contact.merged").latest("created_at")
        self.assertEqual(event.metadata["source_record_preserved"], True)

    def test_merge_blocks_legal_hold_profile_conflict_and_foreign_contact(self):
        primary = Contact.objects.create(
            organization=self.organization,
            first_name="Primary",
            email="same@example.org",
        )
        duplicate = Contact.objects.create(
            organization=self.organization,
            first_name="Duplicate",
            email="same@example.org",
        )
        VolunteerProfile.objects.create(organization=self.organization, contact=primary)
        VolunteerProfile.objects.create(
            organization=self.organization, contact=duplicate
        )
        payload = {
            "primaryContactId": str(primary.id),
            "duplicateContactId": str(duplicate.id),
            "confirm": True,
        }
        conflict = self.client.post(
            "/api/v1/organizations/migration-charity/crm/duplicates/merge/",
            payload,
            format="json",
        )
        self.assertEqual(conflict.status_code, 409)
        VolunteerProfile.objects.filter(contact=duplicate).delete()
        RetentionPolicy.objects.create(
            organization=self.organization,
            record_type="contacts",
            retention_days=365,
            legal_hold=True,
        )
        held = self.client.post(
            "/api/v1/organizations/migration-charity/crm/duplicates/merge/",
            payload,
            format="json",
        )
        self.assertEqual(held.status_code, 423)

        foreign = Contact.objects.create(
            organization=self.foreign_organization, first_name="Foreign"
        )
        foreign_attempt = self.client.post(
            "/api/v1/organizations/migration-charity/crm/duplicates/merge/",
            {
                "primaryContactId": str(primary.id),
                "duplicateContactId": str(foreign.id),
                "confirm": True,
            },
            format="json",
        )
        self.assertEqual(foreign_attempt.status_code, 404)
