from __future__ import annotations

import csv
import hashlib
import io
import re
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import PurePath
from typing import Any

from django.conf import settings
from django.core import signing
from django.core.exceptions import ValidationError as DjangoValidationError
from django.core.validators import validate_email
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from .models import (
    ConsentRecord,
    Contact,
    ContactRelationship,
    DonorSnapshot,
    EmailMessage,
    Household,
    Interaction,
    RetentionPolicy,
    ScheduleEvent,
    VolunteerProfile,
    WaitlistEntry,
)


IMPORT_SCHEMA_VERSION = 1
PREVIEW_SIGNING_SALT = "project-hope.crm-import-preview.v1"
IMPORT_COLUMNS = [
    "contact_type",
    "first_name",
    "last_name",
    "organization_name",
    "preferred_name",
    "email",
    "phone",
    "external_ref",
    "sensitivity",
    "consent_status",
    "notes",
]
EXPORT_COLUMNS = [
    "project_hope_id",
    *IMPORT_COLUMNS,
    "record_status",
    "merged_into_id",
    "created_at",
    "updated_at",
]
READ_ONLY_COLUMNS = {
    "project_hope_id",
    "record_status",
    "merged_into_id",
    "created_at",
    "updated_at",
}


class CRMMigrationError(Exception):
    def __init__(
        self,
        detail: str,
        *,
        status_code: int = 400,
        extra: dict[str, Any] | None = None,
    ):
        super().__init__(detail)
        self.detail = detail
        self.status_code = status_code
        self.extra = extra or {}

    def response_data(self) -> dict[str, Any]:
        return {"detail": self.detail, **self.extra}


@dataclass(frozen=True)
class ParsedContactFile:
    file_name: str
    file_type: str
    digest: str
    headers: list[str]
    rows: list[tuple[int, list[str]]]
    formula_cells: set[tuple[int, int]]
    warnings: list[str]


def _safe_file_name(value: str) -> str:
    name = value.replace("\\", "/").split("/")[-1].strip()
    return (name or "contacts")[:160]


def _read_upload(uploaded_file) -> tuple[bytes, str, str]:
    if uploaded_file is None:
        raise CRMMigrationError("Choose a CSV or XLSX contact file to continue.")
    if uploaded_file.size > settings.PROJECT_HOPE_MAX_CRM_IMPORT_BYTES:
        raise CRMMigrationError(
            "The contact file is larger than the configured import limit.",
            status_code=413,
            extra={"maxBytes": settings.PROJECT_HOPE_MAX_CRM_IMPORT_BYTES},
        )
    data = b"".join(uploaded_file.chunks())
    uploaded_file.seek(0)
    if not data:
        raise CRMMigrationError("The contact file is empty.")
    return data, _safe_file_name(uploaded_file.name), hashlib.sha256(data).hexdigest()


def _validate_xlsx_archive(data: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            members = archive.infolist()
            names = {member.filename for member in members}
            total_size = sum(member.file_size for member in members)
            compressed_size = sum(member.compress_size for member in members)
            unsafe_name = any(
                name.startswith(("/", "\\"))
                or ".." in name.replace("\\", "/").split("/")
                for name in names
            )
            forbidden = any(
                name.lower() == "xl/vbaproject.bin"
                or name.lower().startswith("xl/externallinks/")
                for name in names
            )
            if "[Content_Types].xml" not in names or unsafe_name or forbidden:
                raise ValueError("unsafe workbook archive")
            if len(members) > settings.PROJECT_HOPE_MAX_DOCUMENT_ARCHIVE_MEMBERS:
                raise CRMMigrationError(
                    "The workbook contains too many internal files.", status_code=413
                )
            if total_size > settings.PROJECT_HOPE_MAX_UNCOMPRESSED_DOCUMENT_BYTES:
                raise CRMMigrationError(
                    "The workbook expands beyond the configured safety limit.",
                    status_code=413,
                )
            if compressed_size and total_size > compressed_size * 1000:
                raise CRMMigrationError(
                    "The workbook compression ratio is unsafe.", status_code=413
                )
    except CRMMigrationError:
        raise
    except (OSError, ValueError, zipfile.BadZipFile) as exc:
        raise CRMMigrationError(
            "The XLSX workbook is invalid or contains unsupported linked or macro content."
        ) from exc


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _parse_xlsx(data: bytes, file_name: str, digest: str) -> ParsedContactFile:
    _validate_xlsx_archive(data)
    try:
        workbook = load_workbook(
            io.BytesIO(data), read_only=True, data_only=False, keep_links=False
        )
    except Exception as exc:
        raise CRMMigrationError("The XLSX workbook could not be read safely.") from exc
    try:
        worksheet = (
            workbook["Contacts"]
            if "Contacts" in workbook.sheetnames
            else workbook[workbook.sheetnames[0]]
        )
        iterator = worksheet.iter_rows()
        try:
            header_cells = next(iterator)
        except StopIteration as exc:
            raise CRMMigrationError(
                "The workbook does not contain a header row."
            ) from exc
        headers = [_cell_text(cell.value).strip() for cell in header_cells]
        if len(headers) > 50:
            raise CRMMigrationError("The workbook has more than 50 columns.")
        rows: list[tuple[int, list[str]]] = []
        formula_cells: set[tuple[int, int]] = set()
        for excel_row_number, cells in enumerate(iterator, start=2):
            values = [_cell_text(cell.value) for cell in cells[: len(headers)]]
            if not any(value.strip() for value in values):
                continue
            for column_index, cell in enumerate(cells[: len(headers)]):
                if cell.data_type == "f":
                    formula_cells.add((excel_row_number, column_index))
            rows.append((excel_row_number, values))
            if len(rows) > settings.PROJECT_HOPE_MAX_CRM_IMPORT_ROWS:
                raise CRMMigrationError(
                    "The workbook has more contact rows than this import allows.",
                    status_code=413,
                    extra={"maxRows": settings.PROJECT_HOPE_MAX_CRM_IMPORT_ROWS},
                )
        warnings = []
        if worksheet.title != "Contacts":
            warnings.append(
                f'Imported the first worksheet, "{worksheet.title}". Name it Contacts to make the choice explicit.'
            )
        return ParsedContactFile(
            file_name=file_name,
            file_type="xlsx",
            digest=digest,
            headers=headers,
            rows=rows,
            formula_cells=formula_cells,
            warnings=warnings,
        )
    finally:
        workbook.close()


def _decode_csv(data: bytes) -> tuple[str, list[str]]:
    try:
        return data.decode("utf-8-sig"), []
    except UnicodeDecodeError:
        try:
            return data.decode("cp1252"), [
                "The CSV used Windows-1252 text. Project Hope exports UTF-8 for safer reuse."
            ]
        except UnicodeDecodeError as exc:
            raise CRMMigrationError(
                "The CSV must use UTF-8 or Windows-1252 text."
            ) from exc


def _parse_csv(data: bytes, file_name: str, digest: str) -> ParsedContactFile:
    if b"\x00" in data:
        raise CRMMigrationError("Binary content is not allowed in a CSV contact file.")
    text, warnings = _decode_csv(data)
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    try:
        parsed_rows = csv.reader(io.StringIO(text, newline=""), dialect)
        headers = [value.strip() for value in next(parsed_rows)]
    except (StopIteration, csv.Error) as exc:
        raise CRMMigrationError(
            "The CSV does not contain a readable header row."
        ) from exc
    if len(headers) > 50:
        raise CRMMigrationError("The CSV has more than 50 columns.")
    rows: list[tuple[int, list[str]]] = []
    try:
        for csv_row_number, values in enumerate(parsed_rows, start=2):
            values = [value for value in values[: len(headers)]] + [
                "" for _ in range(max(0, len(headers) - len(values)))
            ]
            if not any(value.strip() for value in values):
                continue
            rows.append((csv_row_number, values))
            if len(rows) > settings.PROJECT_HOPE_MAX_CRM_IMPORT_ROWS:
                raise CRMMigrationError(
                    "The CSV has more contact rows than this import allows.",
                    status_code=413,
                    extra={"maxRows": settings.PROJECT_HOPE_MAX_CRM_IMPORT_ROWS},
                )
    except csv.Error as exc:
        raise CRMMigrationError(
            "The CSV contains malformed quoting or row data."
        ) from exc
    return ParsedContactFile(
        file_name=file_name,
        file_type="csv",
        digest=digest,
        headers=headers,
        rows=rows,
        formula_cells=set(),
        warnings=warnings,
    )


def parse_contact_upload(uploaded_file) -> ParsedContactFile:
    data, file_name, digest = _read_upload(uploaded_file)
    suffix = PurePath(file_name.lower()).suffix
    if suffix == ".xlsx":
        return _parse_xlsx(data, file_name, digest)
    if suffix in {".csv", ".tsv"}:
        return _parse_csv(data, file_name, digest)
    raise CRMMigrationError(
        "Project Hope imports modern XLSX workbooks and CSV/TSV files. Legacy XLS files must be saved as XLSX first.",
        status_code=415,
    )


def _header_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.strip().casefold())


HEADER_ALIASES = {
    _header_key(alias): field
    for field, aliases in {
        "contact_type": ["contact_type", "record type", "type", "contact category"],
        "first_name": ["first_name", "first name", "given name", "firstname"],
        "last_name": ["last_name", "last name", "family name", "surname", "lastname"],
        "organization_name": [
            "organization_name",
            "organization name",
            "organisation",
            "organization",
            "company",
        ],
        "preferred_name": ["preferred_name", "preferred name", "known as"],
        "email": ["email", "email address", "e-mail", "primary email"],
        "phone": ["phone", "phone number", "telephone", "mobile", "primary phone"],
        "external_ref": [
            "external_ref",
            "external reference",
            "legacy id",
            "source id",
            "contact id",
        ],
        "sensitivity": ["sensitivity", "classification", "data classification"],
        "consent_status": [
            "consent_status",
            "consent status",
            "consent",
            "email consent",
        ],
        "notes": ["notes", "note", "comments"],
        "project_hope_id": ["project_hope_id", "project hope id"],
        "record_status": ["record_status", "record status"],
        "merged_into_id": ["merged_into_id", "merged into id"],
        "created_at": ["created_at", "created at"],
        "updated_at": ["updated_at", "updated at"],
    }.items()
    for alias in aliases
}


def _column_mapping(headers: list[str]) -> tuple[dict[int, str], list[str]]:
    mapping: dict[int, str] = {}
    seen: dict[str, str] = {}
    unknown: list[str] = []
    for index, header in enumerate(headers):
        if not header:
            continue
        canonical = HEADER_ALIASES.get(_header_key(header))
        if canonical is None:
            unknown.append(header[:80])
            continue
        if canonical in seen:
            raise CRMMigrationError(
                f'The columns "{seen[canonical]}" and "{header}" both map to {canonical}. Keep only one.'
            )
        mapping[index] = canonical
        seen[canonical] = header
    if not any(field in IMPORT_COLUMNS for field in mapping.values()):
        raise CRMMigrationError(
            "No supported contact columns were found. Download the Project Hope template and copy your data into it."
        )
    warnings = []
    if unknown:
        preview = ", ".join(unknown[:8])
        suffix = " and more" if len(unknown) > 8 else ""
        warnings.append(f"Ignored unrecognized columns: {preview}{suffix}.")
    ignored_readonly = [
        field for field in mapping.values() if field in READ_ONLY_COLUMNS
    ]
    if ignored_readonly:
        warnings.append(
            "Project Hope IDs, merge state, and timestamps are export-only and were not imported."
        )
    return mapping, warnings


VALUE_ALIASES = {
    "contact_type": {
        "individual": "person",
        "individual person": "person",
        "organisation": "organization",
        "organisation contact": "organization",
        "service recipient": "service_user",
        "client": "service_user",
    },
    "sensitivity": {
        "highly sensitive": "highly_sensitive",
        "high sensitivity": "highly_sensitive",
    },
    "consent_status": {
        "yes": "granted",
        "opted in": "granted",
        "opt in": "granted",
        "no": "withdrawn",
        "opted out": "withdrawn",
        "opt out": "withdrawn",
    },
}


def _normalize_import_value(field: str, value: str) -> str:
    clean = value.strip()
    if clean.startswith("'") and len(clean) > 1 and clean[1] in "=+-@":
        clean = clean[1:]
    if field == "email":
        clean = clean.casefold()
    if field in VALUE_ALIASES:
        key = clean.casefold().replace("-", " ").replace("_", " ")
        clean = VALUE_ALIASES[field].get(key, clean.casefold().replace(" ", "_"))
    return clean


def _row_values(
    parsed: ParsedContactFile,
    mapping: dict[int, str],
    row_number: int,
    cells: list[str],
) -> tuple[dict[str, str], list[str]]:
    values: dict[str, str] = {}
    formula_fields: list[str] = []
    for column_index, field in mapping.items():
        if field in READ_ONLY_COLUMNS:
            continue
        value = cells[column_index] if column_index < len(cells) else ""
        values[field] = _normalize_import_value(field, value)
        if (row_number, column_index) in parsed.formula_cells:
            formula_fields.append(field)
    return values, formula_fields


def _validate_row(
    values: dict[str, str], formula_fields: list[str]
) -> dict[str, list[str]]:
    errors: dict[str, list[str]] = defaultdict(list)
    for field in formula_fields:
        errors[field].append(
            "Spreadsheet formulas are not imported. Replace this formula with its reviewed text value."
        )
    choices: dict[str, set[str]] = {
        "contact_type": {choice for choice, _ in Contact.ContactType.choices},
        "sensitivity": {
            choice for choice, _ in Contact._meta.get_field("sensitivity").choices
        },
        "consent_status": {choice for choice, _ in Contact.ConsentStatus.choices},
    }
    defaults: dict[str, str] = {
        "contact_type": "person",
        "sensitivity": "internal",
        "consent_status": "unknown",
    }
    for field, default in defaults.items():
        if not values.get(field):
            values[field] = default
        if values[field] not in choices[field]:
            errors[field].append(f"Use one of: {', '.join(sorted(choices[field]))}.")
    for field in IMPORT_COLUMNS:
        if field not in values:
            continue
        model_field = Contact._meta.get_field(field)
        max_length = getattr(model_field, "max_length", None)
        limit = max_length or (10000 if field == "notes" else None)
        if limit and len(values[field]) > limit:
            errors[field].append(f"Use {limit} characters or fewer.")
    if values.get("email"):
        try:
            validate_email(values["email"])
        except DjangoValidationError:
            errors["email"].append("Enter a valid email address.")
    identity_fields = (
        "first_name",
        "last_name",
        "organization_name",
        "preferred_name",
        "email",
        "phone",
        "external_ref",
    )
    if not any(values.get(field) for field in identity_fields):
        errors["row"].append(
            "Provide a name, organization, email, phone, or external reference."
        )
    if values.get(
        "contact_type"
    ) == Contact.ContactType.ORGANIZATION and not values.get("organization_name"):
        errors["organization_name"].append(
            "Organization records need an organization name."
        )
    return dict(errors)


def _match_key(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().casefold())


def _phone_key(value: str) -> str:
    return "".join(character for character in value if character.isdigit())


def _compatible_contact_types(first_type: str, second_type: str) -> bool:
    return (first_type == Contact.ContactType.ORGANIZATION) == (
        second_type == Contact.ContactType.ORGANIZATION
    )


def _direct_match_reasons(first: Contact, second: Contact) -> list[str]:
    if not _compatible_contact_types(first.contact_type, second.contact_type):
        return []
    reasons: list[str] = []
    if _match_key(first.email) and _match_key(first.email) == _match_key(second.email):
        reasons.append("same email")
    if _match_key(first.external_ref) and _match_key(first.external_ref) == _match_key(
        second.external_ref
    ):
        reasons.append("same external reference")
    first_person_name = (_match_key(first.first_name), _match_key(first.last_name))
    second_person_name = (_match_key(second.first_name), _match_key(second.last_name))
    organization_name_matches = _match_key(first.organization_name) and _match_key(
        first.organization_name
    ) == _match_key(second.organization_name)
    person_name_matches = (
        all(first_person_name) and first_person_name == second_person_name
    )
    if person_name_matches:
        reasons.append("same full name")
    if organization_name_matches:
        reasons.append("same organization name")
    first_display_key = (
        first_person_name
        if all(first_person_name)
        else (_match_key(first.organization_name), "")
    )
    second_display_key = (
        second_person_name
        if all(second_person_name)
        else (_match_key(second.organization_name), "")
    )
    if (
        any(first_display_key)
        and first_display_key == second_display_key
        and _phone_key(first.phone)
        and _phone_key(first.phone) == _phone_key(second.phone)
    ):
        reasons.append("same name and phone")
    return reasons


def _contact_payload(contact: Contact, reasons: list[str]) -> dict[str, Any]:
    return {
        "id": str(contact.id),
        "displayName": contact.display_name,
        "contactType": contact.contact_type,
        "firstName": contact.first_name,
        "lastName": contact.last_name,
        "organizationName": contact.organization_name,
        "email": contact.email,
        "phone": contact.phone,
        "externalRef": contact.external_ref,
        "sensitivity": contact.sensitivity,
        "consentStatus": contact.consent_status,
        "updatedAt": contact.updated_at.isoformat(),
        "matchReasons": reasons,
    }


def _build_contact_indexes(organization):
    contacts = list(
        Contact.objects.filter(
            organization=organization, record_status=Contact.RecordStatus.ACTIVE
        ).order_by("id")
    )
    indexes: dict[str, dict[Any, list[Contact]]] = {
        "email": defaultdict(list),
        "external_ref": defaultdict(list),
        "person_name": defaultdict(list),
        "organization_name": defaultdict(list),
        "name_phone": defaultdict(list),
    }
    for contact in contacts:
        email = _match_key(contact.email)
        external_ref = _match_key(contact.external_ref)
        person_name = (_match_key(contact.first_name), _match_key(contact.last_name))
        organization_name = _match_key(contact.organization_name)
        phone = _phone_key(contact.phone)
        if email:
            indexes["email"][email].append(contact)
        if external_ref:
            indexes["external_ref"][external_ref].append(contact)
        if all(person_name):
            indexes["person_name"][person_name].append(contact)
        if organization_name:
            indexes["organization_name"][organization_name].append(contact)
        display_key = person_name if all(person_name) else (organization_name, "")
        if phone and any(display_key):
            indexes["name_phone"][(display_key, phone)].append(contact)
    return contacts, indexes


def _row_candidates(values: dict[str, str], indexes) -> tuple[list[dict], list[dict]]:
    exact_reasons: dict[str, list[str]] = defaultdict(list)
    possible_reasons: dict[str, list[str]] = defaultdict(list)
    contacts_by_id: dict[str, Contact] = {}

    def add(index_name: str, key: Any, reason: str, exact: bool) -> None:
        if not key:
            return
        for contact in indexes[index_name].get(key, []):
            if not _compatible_contact_types(
                values.get("contact_type", "person"), contact.contact_type
            ):
                continue
            identifier = str(contact.id)
            contacts_by_id[identifier] = contact
            target = exact_reasons if exact else possible_reasons
            target[identifier].append(reason)

    add("email", _match_key(values.get("email", "")), "same email", True)
    add(
        "external_ref",
        _match_key(values.get("external_ref", "")),
        "same external reference",
        True,
    )
    person_name = (
        _match_key(values.get("first_name", "")),
        _match_key(values.get("last_name", "")),
    )
    organization_name = _match_key(values.get("organization_name", ""))
    phone = _phone_key(values.get("phone", ""))
    if all(person_name):
        add("person_name", person_name, "same full name", False)
    if organization_name:
        add(
            "organization_name",
            organization_name,
            "same organization name",
            False,
        )
    display_key = person_name if all(person_name) else (organization_name, "")
    if phone and any(display_key):
        add("name_phone", (display_key, phone), "same name and phone", False)
    for identifier in exact_reasons:
        possible_reasons.pop(identifier, None)
    exact = [
        _contact_payload(contacts_by_id[identifier], reasons)
        for identifier, reasons in exact_reasons.items()
    ]
    possible = [
        _contact_payload(contacts_by_id[identifier], reasons)
        for identifier, reasons in possible_reasons.items()
    ]
    return exact[:10], possible[:10]


def build_contact_preview(uploaded_file, organization, user, *, issue_token=True):
    parsed = parse_contact_upload(uploaded_file)
    mapping, mapping_warnings = _column_mapping(parsed.headers)
    _, indexes = _build_contact_indexes(organization)
    preview_rows: list[dict[str, Any]] = []
    seen_email: dict[str, int] = {}
    seen_external_ref: dict[str, int] = {}
    counts = {
        "totalRows": 0,
        "newRecords": 0,
        "exactMatches": 0,
        "possibleDuplicates": 0,
        "invalidRows": 0,
    }
    for row_number, cells in parsed.rows:
        values, formula_fields = _row_values(parsed, mapping, row_number, cells)
        errors = _validate_row(values, formula_fields)
        email_key = _match_key(values.get("email", ""))
        external_key = _match_key(values.get("external_ref", ""))
        if email_key and email_key in seen_email:
            errors.setdefault("email", []).append(
                f"This email also appears on row {seen_email[email_key]}."
            )
        elif email_key:
            seen_email[email_key] = row_number
        if external_key and external_key in seen_external_ref:
            errors.setdefault("external_ref", []).append(
                f"This external reference also appears on row {seen_external_ref[external_key]}."
            )
        elif external_key:
            seen_external_ref[external_key] = row_number
        exact, possible = _row_candidates(values, indexes)
        if errors:
            row_status = "invalid"
            recommended_action = "skip"
            counts["invalidRows"] += 1
        elif exact:
            row_status = "exact_match"
            recommended_action = "skip"
            counts["exactMatches"] += 1
        elif possible:
            row_status = "possible_duplicate"
            recommended_action = "skip"
            counts["possibleDuplicates"] += 1
        else:
            row_status = "new"
            recommended_action = "create"
            counts["newRecords"] += 1
        provided_fields = [
            field for field in IMPORT_COLUMNS if field in values and values[field] != ""
        ]
        preview_rows.append(
            {
                "rowNumber": row_number,
                "status": row_status,
                "values": values,
                "providedFields": provided_fields,
                "errors": errors,
                "candidates": exact or possible,
                "recommendedAction": recommended_action,
            }
        )
        counts["totalRows"] += 1
    payload = {
        "schemaVersion": IMPORT_SCHEMA_VERSION,
        "fileName": parsed.file_name,
        "fileType": parsed.file_type,
        "fileSha256": parsed.digest,
        "columns": IMPORT_COLUMNS,
        "summary": counts,
        "warnings": [*parsed.warnings, *mapping_warnings],
        "rows": preview_rows,
        "expiresInSeconds": settings.PROJECT_HOPE_CRM_IMPORT_PREVIEW_MAX_AGE_SECONDS,
    }
    if issue_token:
        payload["previewToken"] = signing.dumps(
            {
                "version": IMPORT_SCHEMA_VERSION,
                "organization": str(organization.id),
                "user": str(user.id),
                "digest": parsed.digest,
                "rowCount": counts["totalRows"],
            },
            salt=PREVIEW_SIGNING_SALT,
            compress=True,
        )
    return payload


def validate_preview_token(token: str, preview: dict, organization, user) -> None:
    try:
        signed = signing.loads(
            token,
            salt=PREVIEW_SIGNING_SALT,
            max_age=settings.PROJECT_HOPE_CRM_IMPORT_PREVIEW_MAX_AGE_SECONDS,
        )
    except signing.SignatureExpired as exc:
        raise CRMMigrationError(
            "This import preview expired. Preview the file again before importing.",
            status_code=409,
        ) from exc
    except signing.BadSignature as exc:
        raise CRMMigrationError(
            "The import preview token is invalid.", status_code=400
        ) from exc
    expected = {
        "version": IMPORT_SCHEMA_VERSION,
        "organization": str(organization.id),
        "user": str(user.id),
        "digest": preview["fileSha256"],
        "rowCount": preview["summary"]["totalRows"],
    }
    if signed != expected:
        raise CRMMigrationError(
            "The organization, user, file, or row count changed after preview. Preview the file again.",
            status_code=409,
        )


def _safe_contact_values(row: dict[str, Any]) -> dict[str, str]:
    return {
        field: str(row["values"].get(field, ""))
        for field in IMPORT_COLUMNS
        if field in row["providedFields"]
    }


def _fill_missing_contact_values(contact: Contact, values: dict[str, str]) -> list[str]:
    updated: list[str] = []
    for field in (
        "first_name",
        "last_name",
        "organization_name",
        "preferred_name",
        "email",
        "phone",
        "external_ref",
    ):
        incoming = values.get(field, "")
        if incoming and not getattr(contact, field):
            setattr(contact, field, incoming)
            updated.append(field)
    sensitivity_order = [
        "public",
        "internal",
        "confidential",
        "highly_sensitive",
        "restricted",
    ]
    incoming_sensitivity = values.get("sensitivity")
    if incoming_sensitivity and sensitivity_order.index(
        incoming_sensitivity
    ) > sensitivity_order.index(contact.sensitivity):
        contact.sensitivity = incoming_sensitivity
        updated.append("sensitivity")
    incoming_consent = values.get("consent_status")
    consent_rank = {"unknown": 0, "granted": 1, "withdrawn": 2}
    if (
        incoming_consent
        and consent_rank[incoming_consent] > consent_rank[contact.consent_status]
    ):
        contact.consent_status = incoming_consent
        updated.append("consent_status")
    incoming_notes = values.get("notes", "")
    if incoming_notes and incoming_notes != contact.notes:
        contact.notes = (
            f"{contact.notes}\n\nImported note:\n{incoming_notes}".strip()
            if contact.notes
            else incoming_notes
        )
        updated.append("notes")
    if updated:
        contact.save(update_fields=[*updated, "updated_at"])
    return updated


@transaction.atomic
def commit_contact_import(
    uploaded_file, organization, user, token: str, actions: list[dict]
):
    organization.__class__.objects.select_for_update().get(pk=organization.pk)
    preview = build_contact_preview(
        uploaded_file, organization, user, issue_token=False
    )
    validate_preview_token(token, preview, organization, user)
    rows = {row["rowNumber"]: row for row in preview["rows"]}
    actions_by_row: dict[int, dict] = {}
    for action in actions:
        try:
            raw_row_number = action.get("rowNumber")
            if not isinstance(raw_row_number, (int, str)) or isinstance(
                raw_row_number, bool
            ):
                raise TypeError
            row_number = int(raw_row_number)
        except (TypeError, ValueError) as exc:
            raise CRMMigrationError(
                "Every import action needs a numeric rowNumber."
            ) from exc
        if row_number in actions_by_row or row_number not in rows:
            raise CRMMigrationError(
                "Import actions contain an unknown or repeated row."
            )
        actions_by_row[row_number] = action
    result = {
        "created": 0,
        "updated": 0,
        "unchanged": 0,
        "skipped": 0,
        "invalid": preview["summary"]["invalidRows"],
        "createdIds": [],
        "updatedIds": [],
        "fileSha256": preview["fileSha256"],
    }
    for row_number, row in rows.items():
        requested = actions_by_row.get(row_number, {"action": "skip"})
        action = requested.get("action", "skip")
        if row["status"] == "invalid":
            continue
        if action == "skip":
            result["skipped"] += 1
            continue
        values = _safe_contact_values(row)
        if action == "create":
            if row["status"] != "new":
                raise CRMMigrationError(
                    f"Row {row_number} is no longer new. Preview the file and review its duplicate match.",
                    status_code=409,
                )
            contact = Contact.objects.create(organization=organization, **values)
            result["created"] += 1
            result["createdIds"].append(str(contact.id))
            continue
        if action == "update":
            candidate_ids = {candidate["id"] for candidate in row["candidates"]}
            target_id = str(requested.get("targetContactId", ""))
            if target_id not in candidate_ids:
                raise CRMMigrationError(
                    f"Row {row_number} does not match the selected contact. Preview the file again.",
                    status_code=409,
                )
            contact = (
                Contact.objects.select_for_update()
                .filter(
                    pk=target_id,
                    organization=organization,
                    record_status=Contact.RecordStatus.ACTIVE,
                )
                .first()
            )
            if contact is None:
                raise CRMMigrationError(
                    f"The matched contact for row {row_number} changed or is unavailable.",
                    status_code=409,
                )
            updated_fields = _fill_missing_contact_values(contact, values)
            if updated_fields:
                result["updated"] += 1
                result["updatedIds"].append(str(contact.id))
            else:
                result["unchanged"] += 1
            continue
        raise CRMMigrationError(f'Unsupported action "{action}" for row {row_number}.')
    return result


def _spreadsheet_safe(value: Any, *, csv_output: bool) -> str:
    text = "" if value is None else str(value)
    if csv_output and text.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _contact_export_row(contact: Contact, *, csv_output: bool) -> list[str]:
    values = {
        "project_hope_id": str(contact.id),
        "contact_type": contact.contact_type,
        "first_name": contact.first_name,
        "last_name": contact.last_name,
        "organization_name": contact.organization_name,
        "preferred_name": contact.preferred_name,
        "email": contact.email,
        "phone": contact.phone,
        "external_ref": contact.external_ref,
        "sensitivity": contact.sensitivity,
        "consent_status": contact.consent_status,
        "notes": contact.notes,
        "record_status": contact.record_status,
        "merged_into_id": str(contact.merged_into_id or ""),
        "created_at": contact.created_at.isoformat(),
        "updated_at": contact.updated_at.isoformat(),
    }
    return [
        _spreadsheet_safe(values[column], csv_output=csv_output)
        for column in EXPORT_COLUMNS
    ]


def contacts_csv(contacts, *, template=False) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.writer(output, lineterminator="\r\n")
    writer.writerow(IMPORT_COLUMNS if template else EXPORT_COLUMNS)
    if not template:
        for contact in contacts:
            writer.writerow(_contact_export_row(contact, csv_output=True))
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def contacts_xlsx(contacts, *, template=False) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Contacts"
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    headers = IMPORT_COLUMNS if template else EXPORT_COLUMNS
    worksheet.append(headers)
    header_fill = PatternFill("solid", fgColor="214C3F")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    if not template:
        for contact in contacts:
            worksheet.append(_contact_export_row(contact, csv_output=False))
            for cell in worksheet[worksheet.max_row]:
                cell.data_type = "s"
    widths = {
        "A": 38,
        "B": 18,
        "C": 18,
        "D": 18,
        "E": 28,
        "F": 18,
        "G": 30,
        "H": 20,
        "I": 20,
        "J": 20,
        "K": 20,
        "L": 42,
        "M": 18,
        "N": 38,
        "O": 28,
        "P": 28,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    worksheet.auto_filter.ref = (
        f"A1:{worksheet.cell(1, len(headers)).column_letter}{max(1, worksheet.max_row)}"
    )
    if worksheet.max_row > 1:
        table = Table(
            displayName="ProjectHopeContacts",
            ref=f"A1:{worksheet.cell(1, len(headers)).column_letter}{worksheet.max_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
        )
        worksheet.add_table(table)
    if template:
        choice_lists = {
            "A": [choice for choice, _ in Contact.ContactType.choices],
            "I": [
                choice for choice, _ in Contact._meta.get_field("sensitivity").choices
            ],
            "J": [choice for choice, _ in Contact.ConsentStatus.choices],
        }
        for column, choices in choice_lists.items():
            validation = DataValidation(
                type="list", formula1='"' + ",".join(choices) + '"', allow_blank=True
            )
            worksheet.add_data_validation(validation)
            validation.add(f"{column}2:{column}2501")
    instructions = workbook.create_sheet("Read me")
    instructions.sheet_view.showGridLines = False
    instructions["A1"] = "Project Hope contact migration"
    instructions["A1"].font = Font(size=18, bold=True, color="214C3F")
    guidance = [
        "Keep the Contacts header row unchanged. Common alternatives such as First Name and Phone Number are also recognized.",
        "Each row needs a name, organization, email, phone, or external reference.",
        "Preview never changes records. Review every invalid or possible-duplicate row before importing.",
        "Update fills missing details only; it does not erase an existing value.",
        "Spreadsheet formulas, macros, and external workbook links are not imported.",
        "Allowed record types: person, organization, service_user, donor, volunteer.",
        "Allowed sensitivity values: public, internal, confidential, highly_sensitive, restricted.",
        "Allowed consent values: unknown, granted, withdrawn.",
    ]
    for row_index, text in enumerate(guidance, start=3):
        instructions.cell(row_index, 1, f"{row_index - 2}.")
        instructions.cell(row_index, 2, text)
        instructions.cell(row_index, 2).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
    instructions.column_dimensions["A"].width = 6
    instructions.column_dimensions["B"].width = 100
    workbook.properties.creator = "Project Hope"
    workbook.properties.title = (
        "Project Hope contact migration template"
        if template
        else "Project Hope contact export"
    )
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def duplicate_contact_pairs(organization, *, limit=100) -> dict[str, Any]:
    contacts, indexes = _build_contact_indexes(organization)
    pairs: dict[tuple[str, str], dict[str, Any]] = {}

    def add_groups(groups, reason: str, confidence: str) -> None:
        for grouped_contacts in groups.values():
            if len(grouped_contacts) < 2 or len(grouped_contacts) > 25:
                continue
            for index, first in enumerate(grouped_contacts):
                for second in grouped_contacts[index + 1 :]:
                    if not _compatible_contact_types(
                        first.contact_type, second.contact_type
                    ):
                        continue
                    first_id, second_id = sorted((str(first.id), str(second.id)))
                    key = (first_id, second_id)
                    pair = pairs.setdefault(
                        key,
                        {
                            "first": _contact_payload(first, []),
                            "second": _contact_payload(second, []),
                            "matchReasons": [],
                            "confidence": confidence,
                        },
                    )
                    if reason not in pair["matchReasons"]:
                        pair["matchReasons"].append(reason)
                    if confidence == "exact":
                        pair["confidence"] = "exact"

    add_groups(indexes["email"], "same email", "exact")
    add_groups(indexes["external_ref"], "same external reference", "exact")
    add_groups(indexes["name_phone"], "same name and phone", "strong")
    add_groups(indexes["person_name"], "same full name", "possible")
    add_groups(indexes["organization_name"], "same organization name", "possible")
    ordered = sorted(
        pairs.values(),
        key=lambda item: (
            {"exact": 0, "strong": 1, "possible": 2}[item["confidence"]],
            item["first"]["displayName"].casefold(),
        ),
    )
    return {
        "totalActiveContacts": len(contacts),
        "totalCandidates": len(ordered),
        "results": ordered[: max(1, min(limit, 250))],
    }


def _merge_notes(primary: str, duplicate: str) -> str:
    if not duplicate or duplicate == primary:
        return primary
    if not primary:
        return duplicate
    return f"{primary}\n\nMerged record note:\n{duplicate}"


@transaction.atomic
def merge_contacts(organization, user, primary_id, duplicate_id):
    if str(primary_id) == str(duplicate_id):
        raise CRMMigrationError("Choose two different contacts to merge.")
    contacts = {
        str(contact.id): contact
        for contact in Contact.objects.select_for_update().filter(
            organization=organization,
            id__in=[primary_id, duplicate_id],
            record_status=Contact.RecordStatus.ACTIVE,
        )
    }
    primary = contacts.get(str(primary_id))
    duplicate = contacts.get(str(duplicate_id))
    if primary is None or duplicate is None:
        raise CRMMigrationError(
            "Both contacts must be active records in this organization.",
            status_code=404,
        )
    if not _direct_match_reasons(primary, duplicate):
        raise CRMMigrationError(
            "These contacts no longer have a duplicate match. Review duplicate candidates again before merging.",
            status_code=409,
        )
    if RetentionPolicy.objects.filter(
        organization=organization,
        record_type="contacts",
        enabled=True,
        legal_hold=True,
    ).exists():
        raise CRMMigrationError(
            "Contacts are under legal hold and cannot be merged.", status_code=423
        )
    primary_profile = VolunteerProfile.objects.filter(
        organization=organization, contact=primary
    ).first()
    duplicate_profile = VolunteerProfile.objects.filter(
        organization=organization, contact=duplicate
    ).first()
    if primary_profile and duplicate_profile:
        raise CRMMigrationError(
            "Both contacts have volunteer profiles. Resolve those profiles before merging the contacts.",
            status_code=409,
        )
    updated_fields: list[str] = []
    for field in (
        "first_name",
        "last_name",
        "organization_name",
        "preferred_name",
        "email",
        "phone",
        "external_ref",
    ):
        if not getattr(primary, field) and getattr(duplicate, field):
            setattr(primary, field, getattr(duplicate, field))
            updated_fields.append(field)
    merged_notes = _merge_notes(primary.notes, duplicate.notes)
    if merged_notes != primary.notes:
        primary.notes = merged_notes
        updated_fields.append("notes")
    sensitivity_order = [
        "public",
        "internal",
        "confidential",
        "highly_sensitive",
        "restricted",
    ]
    if sensitivity_order.index(duplicate.sensitivity) > sensitivity_order.index(
        primary.sensitivity
    ):
        primary.sensitivity = duplicate.sensitivity
        updated_fields.append("sensitivity")
    consent_rank = {"unknown": 0, "granted": 1, "withdrawn": 2}
    if consent_rank[duplicate.consent_status] > consent_rank[primary.consent_status]:
        primary.consent_status = duplicate.consent_status
        updated_fields.append("consent_status")
    if updated_fields:
        primary.save(update_fields=[*updated_fields, "updated_at"])

    reassigned = {
        "households": Household.objects.filter(
            organization=organization, primary_contact=duplicate
        ).update(primary_contact=primary),
        "interactions": Interaction.objects.filter(
            organization=organization, contact=duplicate
        ).update(contact=primary),
        "consents": ConsentRecord.objects.filter(
            organization=organization, contact=duplicate
        ).update(contact=primary),
        "schedules": ScheduleEvent.objects.filter(
            organization=organization, contact=duplicate
        ).update(contact=primary),
        "waitlist": WaitlistEntry.objects.filter(
            organization=organization, contact=duplicate
        ).update(contact=primary),
        "emailMessages": EmailMessage.objects.filter(
            organization=organization, crm_contact=duplicate
        ).update(crm_contact=primary),
        "donorSnapshots": DonorSnapshot.objects.filter(
            organization=organization, contact=duplicate
        ).update(contact=primary),
        "volunteerProfiles": 0,
        "relationships": 0,
        "relationshipConflictsCombined": 0,
        "selfRelationshipsPreserved": 0,
    }
    if duplicate_profile and not primary_profile:
        duplicate_profile.contact = primary
        duplicate_profile.save(update_fields=["contact", "updated_at"])
        reassigned["volunteerProfiles"] = 1

    relationships = list(
        ContactRelationship.objects.select_for_update().filter(
            Q(from_contact=duplicate) | Q(to_contact=duplicate),
            organization=organization,
        )
    )
    for relationship in relationships:
        new_from = (
            primary
            if relationship.from_contact_id == duplicate.id
            else relationship.from_contact
        )
        new_to = (
            primary
            if relationship.to_contact_id == duplicate.id
            else relationship.to_contact
        )
        if new_from.id == new_to.id:
            reassigned["selfRelationshipsPreserved"] += 1
            continue
        existing = (
            ContactRelationship.objects.filter(
                organization=organization,
                from_contact=new_from,
                to_contact=new_to,
                relation_type=relationship.relation_type,
            )
            .exclude(pk=relationship.pk)
            .first()
        )
        if existing:
            combined_notes = _merge_notes(existing.notes, relationship.notes)
            if combined_notes != existing.notes:
                existing.notes = combined_notes
                existing.save(update_fields=["notes", "updated_at"])
            relationship.delete()
            reassigned["relationshipConflictsCombined"] += 1
        else:
            relationship.from_contact = new_from
            relationship.to_contact = new_to
            relationship.save(
                update_fields=["from_contact", "to_contact", "updated_at"]
            )
            reassigned["relationships"] += 1

    Contact.objects.filter(organization=organization, merged_into=duplicate).update(
        merged_into=primary
    )
    duplicate.record_status = Contact.RecordStatus.MERGED
    duplicate.merged_into = primary
    duplicate.merged_at = timezone.now()
    duplicate.merged_by = user
    duplicate.save(
        update_fields=[
            "record_status",
            "merged_into",
            "merged_at",
            "merged_by",
            "updated_at",
        ]
    )
    return {
        "primary": _contact_payload(primary, []),
        "mergedContactId": str(duplicate.id),
        "reassigned": reassigned,
        "preserved": True,
    }
